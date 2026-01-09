import pandas as pd
import numpy as np
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
import asyncio
import io
import re
from collections import defaultdict
import warnings
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Конфигурация бота
TOKEN = '8234604599:AAFluxkjpDxUDz2kgvNYTwGIkMi5NuWrfmU'

# Константы для ABC-анализа (80/15/5)
ABC_CATEGORIES = {
    'A': {'min_percent': 0, 'max_percent': 80, 'description': 'Критически важные товары'},
    'B': {'min_percent': 80, 'max_percent': 95, 'description': 'Средняя значимость'},
    'C': {'min_percent': 95, 'max_percent': 100, 'description': 'Наименьшая значимость'}
}

# Цвета для категорий ABC
ABC_COLORS = {
    'A': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Светло-зеленый
    'B': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Светло-желтый
    'C': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Светло-красный
}

# Шрифт для категорий
CATEGORY_FONT = Font(color='000000', bold=True)  # Черный жирный шрифт

# Улучшенные маппинги для столбцов на основе опыта всех версий бота
COLUMN_MAPPINGS = {
    'товар': ['товар', 'наименование', 'название', 'name', 'product', 'артикул', 'код', 'элемент номенклатуры', 'продукт', 'название товара', 'позиция', 'позиция.', 'название позиции', 'наименование товара'],
    'факт': ['факт', 'фактическое', 'факт.', 'actual', 'fact', 'кол-во', 'количество', 'факт кол-во', 'фактическое количество', 'факт. количество', 'фактический остаток'],
    'учет': ['учет', 'книжн', 'книжное', 'бухг', 'учетное', 'book', 'account', 'баланс', 'план', 'учетное количество', 'учетный остаток', 'книжный остаток'],
    'разница': ['разница', 'дельта', 'diff', 'difference', 'delta', 'отклонение'],
    'количество': ['количество', 'quantity', 'qty', 'кол-во', 'объем', 'продажи', 'ед.', 'единицы', 'количество продаж', 'продано', 'штук', 'шт', 'кол', 'кол.', 'колво', 'кол-во продаж', 'количество (шт)'],
    'выручка': ['выручка', 'revenue', 'продажи', 'sales', 'выручка, р.', 'выручка (руб)', 'sum', 'выручка руб', 'выручка, руб', 'выручка (р)', 'выручка(руб)'],
    'прибыль': ['прибыль', 'profit', 'маржа', 'валовая прибыль', 'чистая прибыль', 'маржинальность', 'валовая прибыль, р.', 'прибыль, р.', 'gross profit', 'прибыль руб', 'прибыль,руб', 'прибыль (руб)', 'прибыль(руб)'],
    'наценка': ['наценка', 'markup', 'рентабельность']
}

def normalize_column_name(name):
    """Быстрая нормализация названий столбцов"""
    if not isinstance(name, str):
        name = str(name)
    
    name_lower = name.lower().strip()
    
    for std_name, variants in COLUMN_MAPPINGS.items():
        for variant in variants:
            if variant == name_lower or variant in name_lower:
                return std_name
    return name_lower

NUMBER_CACHE = {}

def parse_number_cached(value):
    """Кешированная функция парсинга чисел"""
    if pd.isna(value):
        return 0.0
    
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    
    if isinstance(value, str):
        if value in NUMBER_CACHE:
            return NUMBER_CACHE[value]
        
        val = value.strip()
        if not val:
            result = 0.0
        else:
            # Заменяем запятые на точки и убираем пробелы
            val = val.replace(',', '.').replace(' ', '')
            
            # Убираем символы процента и другие нечисловые символы, кроме минуса и точки
            val = re.sub(r'[^\d\.\-]', '', val)
            
            try:
                result = float(val)
            except ValueError:
                result = 0.0
        
        NUMBER_CACHE[value] = result
        return result
    
    return 0.0

# Функция ABC-анализа с улучшенной обработкой
def perform_abc_analysis(file_content):
    """Выполняет ABC-анализ по файлу продаж"""
    try:
        excel_file = pd.ExcelFile(io.BytesIO(file_content), engine='openpyxl')
        
        all_abc_results = {}
        processed_sheets = []
        skipped_sheets = []
        
        for sheet_name in excel_file.sheet_names:
            try:
                logger.info(f"Обработка листа: {sheet_name}")
                
                # Читаем лист как есть
                df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str, header=None)
                
                # Пробуем найти строку заголовка
                header_row = 0
                for i in range(min(5, len(df))):  # Проверяем первые 5 строк
                    row_values = df.iloc[i].astype(str).fillna('').tolist()
                    row_str = ' '.join(row_values).lower()
                    # Ищем ключевые слова в строке
                    if any(keyword in row_str for keyword in ['товар', 'наименование', 'название', 'product', 'артикул', 'код', 'продукт', 'позиция']):
                        header_row = i
                        break
                
                # Перечитываем с правильной строкой заголовка
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row, dtype=str)
                
                if df.empty or len(df.columns) < 2:
                    logger.warning(f"Лист '{sheet_name}' пустой или мало столбцов")
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Пустой лист или мало данных'})
                    continue
                
                # Нормализуем названия столбцов
                original_columns = df.columns.tolist()
                df.columns = [normalize_column_name(col) for col in df.columns]
                
                logger.info(f"Лист '{sheet_name}' столбцы после нормализации: {list(df.columns)}")
                
                # Проверяем наличие столбца с товарами
                if 'товар' not in df.columns:
                    # Пытаемся найти столбец с товаром по содержимому
                    for col in df.columns:
                        sample_values = df[col].dropna().head(10).astype(str).tolist()
                        # Более гибкая проверка на товары
                        sample_str = ' '.join(sample_values).lower()
                        if any(keyword in sample_str for keyword in ['пицца', 'бургер', 'латте', 'кофе', 'салат', 'суп', 'напиток', 'десерт', 'сэндвич', 'капучино', 'американо', 'пончик', 'вафли', 'сок', 'вода', 'панини', 'сыр', 'брут', 'круасс', 'торт']):
                            df = df.rename(columns={col: 'товар'})
                            logger.info(f"Автоопределен столбец товара: {col} -> товар")
                            break
                        # Если в столбце есть уникальные названия (не числа)
                        if len(set(sample_values)) > len(sample_values) * 0.5 and len(sample_values) > 3:
                            df = df.rename(columns={col: 'товар'})
                            logger.info(f"Автоопределен столбец товара (уникальные значения): {col} -> товар")
                            break
                
                if 'товар' not in df.columns:
                    logger.warning(f"На листе '{sheet_name}' не найден столбец товара")
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Не найден столбец товара'})
                    continue
                
                # Ищем столбцы прибыли и количества
                profit_column = None
                quantity_column = None
                
                for col in df.columns:
                    col_lower = col.lower()
                    if col_lower == 'прибыль' or col_lower.startswith('прибыль') or 'profit' in col_lower:
                        profit_column = col
                        logger.info(f"Найден столбец прибыли: {col}")
                    elif col_lower == 'количество' or col_lower.startswith('количество') or 'qty' in col_lower or 'кол-во' in col_lower or 'шт' in col_lower:
                        quantity_column = col
                        logger.info(f"Найден столбец количества: {col}")
                
                # Если не нашли прибыль, ищем числовые столбцы
                if not profit_column:
                    numeric_cols = []
                    for col in df.columns:
                        if col != 'товар':
                            try:
                                sample = df[col].dropna().head(20)
                                if len(sample) > 0:
                                    numeric_count = 0
                                    total_sum = 0
                                    for val in sample:
                                        parsed = parse_number_cached(val)
                                        if parsed > 0:
                                            numeric_count += 1
                                            total_sum += parsed
                                    
                                    if numeric_count > 0:
                                        avg_value = total_sum / numeric_count
                                        if avg_value > 50:  # Прибыль обычно имеет бОльшие значения
                                            numeric_cols.append((col, numeric_count, avg_value))
                            except:
                                pass
                    
                    if numeric_cols:
                        numeric_cols.sort(key=lambda x: x[2], reverse=True)
                        profit_column = numeric_cols[0][0]
                        logger.info(f"Автоопределен столбец прибыли: {profit_column} (среднее: {numeric_cols[0][2]:.2f})")
                
                if not profit_column:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Не найден столбец прибыли'})
                    continue
                
                # Ищем количество, если не нашли
                if not quantity_column:
                    for col in df.columns:
                        if col not in ['товар', profit_column]:
                            sample = df[col].dropna().head(10)
                            if len(sample) > 0:
                                numeric_vals = []
                                for val in sample:
                                    parsed = parse_number_cached(val)
                                    if parsed > 0:
                                        numeric_vals.append(parsed)
                                
                                if len(numeric_vals) > 5:
                                    avg_qty = sum(numeric_vals) / len(numeric_vals)
                                    # Количество обычно целые числа и не слишком большие
                                    if 0.5 < avg_qty < 1000 and all(v == int(v) for v in numeric_vals[:5]):
                                        quantity_column = col
                                        logger.info(f"Автоопределен столбец количества: {col} (среднее: {avg_qty:.2f})")
                                        break
                
                # Выбираем нужные столбцы
                columns_to_use = ['товар', profit_column]
                if quantity_column:
                    columns_to_use.append(quantity_column)
                
                # Создаем чистую копию
                df_clean = df[columns_to_use].copy()
                
                # Очищаем данные
                df_clean = df_clean.dropna(subset=['товар'])
                df_clean['товар'] = df_clean['товар'].astype(str).str.strip()
                
                # Убираем строки с итогами
                summary_keywords = ['итого', 'total', 'всего', 'сумма', 'итог', 'общий', 'total:', 'итого:', 'всего:']
                df_clean = df_clean[~df_clean['товар'].str.lower().isin([kw.lower() for kw in summary_keywords])]
                df_clean = df_clean[df_clean['товар'] != '']
                
                if len(df_clean) == 0:
                    logger.warning(f"На листе '{sheet_name}' нет товаров после очистки")
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Нет данных после очистки'})
                    continue
                
                # Парсим прибыль
                df_clean['прибыль_число'] = df_clean[profit_column].apply(parse_number_cached)
                
                # Убираем нулевые и отрицательные значения прибыли
                df_clean = df_clean[df_clean['прибыль_число'] > 0]
                
                if len(df_clean) == 0:
                    logger.warning(f"На листе '{sheet_name}' нет товаров с положительной прибылью")
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Нет положительной прибыли'})
                    continue
                
                # Парсим количество
                if quantity_column:
                    df_clean['количество_число'] = df_clean[quantity_column].apply(parse_number_cached)
                    # Проверяем, не все ли значения 1
                    unique_values = df_clean['количество_число'].unique()
                    if len(unique_values) == 1 and unique_values[0] == 1:
                        logger.warning(f"На листе '{sheet_name}' все значения количества равны 1")
                        # Ищем другие потенциальные столбцы количества
                        for col in df.columns:
                            if col not in columns_to_use:
                                try:
                                    sample = df[col].dropna().head(20)
                                    if len(sample) > 0:
                                        numeric_vals = []
                                        for val in sample:
                                            parsed = parse_number_cached(val)
                                            if parsed > 0:
                                                numeric_vals.append(parsed)
                                        
                                        if len(numeric_vals) > 5:
                                            avg_val = sum(numeric_vals) / len(numeric_vals)
                                            if avg_val > 1 and avg_val < 100:
                                                df_clean['количество_число'] = df[col].apply(parse_number_cached)
                                                quantity_column = col
                                                logger.info(f"Найден альтернативный столбец количества: {col}")
                                                break
                                except:
                                    pass
                else:
                    df_clean['количество_число'] = 1
                    logger.info(f"На листе '{sheet_name}' столбец количества не найден, используется значение 1")
                
                # ABC-анализ по прибыли
                df_sorted_profit = df_clean.sort_values('прибыль_число', ascending=False).reset_index(drop=True)
                total_profit = df_sorted_profit['прибыль_число'].sum()
                
                df_sorted_profit['доля_в_прибыли'] = (df_sorted_profit['прибыль_число'] / total_profit) * 100
                df_sorted_profit['кумулятивная_доля_прибыли'] = df_sorted_profit['доля_в_прибыли'].cumsum()
                
                # Присваиваем категории ABC
                def assign_abc_category(cumulative_share):
                    if cumulative_share <= 80:
                        return 'A'
                    elif cumulative_share <= 95:
                        return 'B'
                    else:
                        return 'C'
                
                df_sorted_profit['категория_abc_прибыль'] = df_sorted_profit['кумулятивная_доля_прибыли'].apply(assign_abc_category)
                
                # ABC-анализ по количеству
                total_qty = df_sorted_profit['количество_число'].sum()
                df_sorted_profit['доля_в_количестве'] = (df_sorted_profit['количество_число'] / total_qty) * 100
                df_sorted_profit['кумулятивная_доля_количества'] = df_sorted_profit['доля_в_количестве'].cumsum()
                df_sorted_profit['категория_abc_количество'] = df_sorted_profit['кумулятивная_доля_количества'].apply(assign_abc_category)
                
                # Добавляем позицию
                df_sorted_profit['позиция_прибыль'] = df_sorted_profit.index + 1
                
                # Статистика по категориям
                category_stats = {}
                for category in ['A', 'B', 'C']:
                    cat_data = df_sorted_profit[df_sorted_profit['категория_abc_прибыль'] == category]
                    if len(cat_data) > 0:
                        category_stats[category] = {
                            'количество_товаров': len(cat_data),
                            'доля_товаров': (len(cat_data) / len(df_sorted_profit)) * 100,
                            'сумма_прибыли': cat_data['прибыль_число'].sum(),
                            'доля_прибыли': cat_data['доля_в_прибыли'].sum(),
                            'средняя_прибыль': cat_data['прибыль_число'].mean(),
                            'сумма_количества': cat_data['количество_число'].sum(),
                            'среднее_количество': cat_data['количество_число'].mean()
                        }
                    else:
                        category_stats[category] = {
                            'количество_товаров': 0,
                            'доля_товаров': 0,
                            'сумма_прибыли': 0,
                            'доля_прибыли': 0,
                            'средняя_прибыль': 0,
                            'сумма_количества': 0,
                            'среднее_количество': 0
                        }
                
                # Топ товаров категории A
                top_a_products = df_sorted_profit[df_sorted_profit['категория_abc_прибыль'] == 'A'][['товар', 'прибыль_число', 'доля_в_прибыли', 'количество_число']].head(10)
                
                all_abc_results[sheet_name] = {
                    'dataframe': df_sorted_profit,
                    'total_profit': total_profit,
                    'total_quantity': total_qty,
                    'total_products': len(df_sorted_profit),
                    'category_stats': category_stats,
                    'top_a_products': top_a_products,
                    'profit_column': profit_column,
                    'quantity_column': quantity_column if quantity_column else 'не найден'
                }
                
                processed_sheets.append(sheet_name)
                logger.info(f"Успешно обработан лист '{sheet_name}': {len(df_sorted_profit)} товаров, прибыль: {total_profit:.2f}, количество: {total_qty:.0f}")
                
            except Exception as e:
                logger.error(f"Ошибка в листе {sheet_name}: {str(e)}", exc_info=True)
                skipped_sheets.append({'sheet': sheet_name, 'reason': f'Ошибка обработки: {str(e)[:100]}'})
                continue
        
        NUMBER_CACHE.clear()
        
        # Добавляем информацию о пропущенных листах
        if all_abc_results:
            for sheet_name in all_abc_results.keys():
                all_abc_results[sheet_name]['processed_sheets'] = processed_sheets
                all_abc_results[sheet_name]['skipped_sheets'] = skipped_sheets
        
        return all_abc_results if all_abc_results else None
        
    except Exception as e:
        logger.error(f"Ошибка ABC-анализа: {str(e)}", exc_info=True)
        NUMBER_CACHE.clear()
        return None

# Улучшенная функция создания отчета с настройкой размеров
def create_abc_excel_report(abc_results, original_filename):
    """Создает Excel файл с улучшенными рекомендациями и цветовым кодированием"""
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Сводный отчет
        summary_data = []
        for warehouse, data in abc_results.items():
            stats = data['category_stats']
            summary_data.append({
                'Склад/Точка': warehouse,
                'Всего товаров': data['total_products'],
                'Общая прибыль': data['total_profit'],
                'Общее количество': data['total_quantity'],
                'Товаров A': stats['A']['количество_товаров'],
                'Доля A, %': f"{stats['A']['доля_прибыли']:.1f}%",
                'Товаров B': stats['B']['количество_товаров'],
                'Доля B, %': f"{stats['B']['доля_прибыли']:.1f}%",
                'Товаров C': stats['C']['количество_товаров'],
                'Доля C, %': f"{stats['C']['доля_прибыли']:.1f}%"
            })
        
        summary_df = pd.DataFrame(summary_data)
        # Форматируем числовые столбцы
        summary_df['Общая прибыль'] = summary_df['Общая прибыль'].apply(lambda x: f"{x:,.2f} р.")
        summary_df['Общее количество'] = summary_df['Общее количество'].apply(lambda x: f"{x:,.0f}")
        
        summary_df.to_excel(writer, sheet_name='Сводка', index=False)
        
        # Настраиваем ширину столбцов для сводки
        worksheet = writer.sheets['Сводка']
        column_widths = {
            'A': 30,  # Склад/Точка
            'B': 15,  # Всего товаров
            'C': 20,  # Общая прибыль
            'D': 20,  # Общее количество
            'E': 15,  # Товаров A
            'F': 15,  # Доля A, %
            'G': 15,  # Товаров B
            'H': 15,  # Доля B, %
            'I': 15,  # Товаров C
            'J': 15   # Доля C, %
        }
        for col, width in column_widths.items():
            if col in worksheet.column_dimensions:
                worksheet.column_dimensions[col].width = width
        
        # 2. Детальный анализ по складам с цветовым кодированием
        for warehouse, data in abc_results.items():
            df = data['dataframe'].copy()
            
            report_df = df[[
                'позиция_прибыль', 'товар', 'прибыль_число', 'доля_в_прибыли', 
                'кумулятивная_доля_прибыли', 'категория_abc_прибыль',
                'количество_число', 'доля_в_количестве', 'кумулятивная_доля_количества', 'категория_abc_количество'
            ]].copy()
            
            report_df.columns = [
                'Позиция', 'Товар', 'Прибыль, руб.', 'Доля в прибыли, %',
                'Кумулятивная доля прибыли, %', 'Категория ABC (прибыль)',
                'Количество продаж', 'Доля в количестве, %',
                'Кумулятивная доля количества, %', 'Категория ABC (количество)'
            ]
            
            # Форматируем числа
            report_df['Прибыль, руб.'] = report_df['Прибыль, руб.'].apply(lambda x: f"{x:,.2f}")
            report_df['Доля в прибыли, %'] = report_df['Доля в прибыли, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Кумулятивная доля прибыли, %'] = report_df['Кумулятивная доля прибыли, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Количество продаж'] = report_df['Количество продаж'].apply(lambda x: f"{x:,.0f}")
            report_df['Доля в количестве, %'] = report_df['Доля в количестве, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Кумулятивная доля количества, %'] = report_df['Кумулятивная доля количества, %'].apply(lambda x: f"{x:.2f}%")
            
            sheet_name = f"ABC_{warehouse[:20]}" if len(warehouse) > 20 else f"ABC_{warehouse}"
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', sheet_name)
            
            report_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Настраиваем ширину столбцов и применяем цветовое кодирование
            if sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Автоматическая настройка ширины столбцов
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Ручная настройка ключевых столбцов
                worksheet.column_dimensions['B'].width = 35  # Товар
                worksheet.column_dimensions['F'].width = 22  # Категория ABC (прибыль)
                worksheet.column_dimensions['J'].width = 22  # Категория ABC (количество)
                
                # Применяем цветовое кодирование для категорий ABC
                for row in range(2, len(report_df) + 2):  # Начинаем с строки 2 (после заголовка)
                    # Категория ABC по прибыли (столбец F)
                    profit_cell = worksheet[f'F{row}']
                    category = profit_cell.value
                    if category in ABC_COLORS:
                        profit_cell.fill = ABC_COLORS[category]
                        profit_cell.font = CATEGORY_FONT
                        profit_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Категория ABC по количеству (столбец J)
                    qty_cell = worksheet[f'J{row}']
                    category_qty = qty_cell.value
                    if category_qty in ABC_COLORS:
                        qty_cell.fill = ABC_COLORS[category_qty]
                        qty_cell.font = CATEGORY_FONT
                        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 3. Расширенные рекомендации с анализом проблем
        recommendations = [
            {
                'Категория': 'A',
                'Доля прибыли': '80%',
                'Доля товаров': '~20%',
                'Стратегия управления': 'Тщательный контроль на запасов в Частый монитор',
                'Рекомендации по оптимизации': 'Увеличить запас на 10-15%',
                'Контрольные показатели': 'Уровень сервиса > 98%'
            },
            {
                'Категория': 'B',
                'Доля прибыли': '15%',
                'Доля товаров': '~30%',
                'Стратегия управления': 'Умеренный контроль',
                'Рекомендации по оптимизации': 'Оптимизировать партии заказа ППАВС-анализ',
                'Контрольные показатели': 'Уровень сервиса > 95%'
            },
            {
                'Категория': 'C',
                'Доля прибыли': '5%',
                'Доля товаров': '~50%',
                'Стратегия управления': 'Периодические проверки',
                'Рекомендации по оптимизации': 'Снизить страховой запас и объединить с другими позициями',
                'Контрольные показатели': 'Уровень сервиса > 90%'
            }
        ]
        
        rec_df = pd.DataFrame(recommendations)
        rec_df.to_excel(writer, sheet_name='Рекомендации_управления', index=False)
        
        worksheet = writer.sheets['Рекомендации_управления']
        
        # Настраиваем ширину столбцов
        column_widths = {
            'A': 15,  # Категория
            'B': 15,  # Доля прибыли
            'C': 15,  # Доля товаров
            'D': 35,  # Стратегия управления
            'E': 40,  # Рекомендации по оптимизации
            'F': 30   # Контрольные показатели
        }
        for col, width in column_widths.items():
            if col in worksheet.column_dimensions:
                worksheet.column_dimensions[col].width = width
        
        # Включаем перенос текста и выравнивание
        wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        for row in range(2, 5):  # Строки 2-4 (данные)
            # Устанавливаем высоту строк
            worksheet.row_dimensions[row].height = 70
            
            # Применяем цветовое кодирование для категорий
            category_cell = worksheet[f'A{row}']
            category = category_cell.value
            if category in ABC_COLORS:
                category_cell.fill = ABC_COLORS[category]
                category_cell.font = CATEGORY_FONT
                category_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Применяем перенос текста ко всем ячейкам с текстом
            for col in ['D', 'E', 'F']:
                cell = worksheet[f'{col}{row}']
                cell.alignment = wrap_alignment
        
        # 4. Топ товаров
        top_a_all = []
        for warehouse, data in abc_results.items():
            top_a = data['top_a_products']
            for _, row in top_a.iterrows():
                top_a_all.append({
                    'Склад/Точка': warehouse,
                    'Товар': row['товар'][:50],
                    'Прибыль, руб.': row['прибыль_число'],
                    'Доля, %': row['доля_в_прибыли'],
                    'Количество': row['количество_число']
                })
        
        if top_a_all:
            top_a_df = pd.DataFrame(top_a_all)
            top_a_df = top_a_df.sort_values('Прибыль, руб.', ascending=False)
            top_a_df['Прибыль, руб.'] = top_a_df['Прибыль, руб.'].apply(lambda x: f"{x:,.2f}")
            top_a_df['Доля, %'] = top_a_df['Доля, %'].apply(lambda x: f"{x:.2f}%")
            top_a_df['Количество'] = top_a_df['Количество'].apply(lambda x: f"{x:,.0f}")
            top_a_df.to_excel(writer, sheet_name='Топ_A_товары', index=False)
            
            worksheet = writer.sheets['Топ_A_товары']
            
            # Автоматическая настройка ширины
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 5. Пропущенные листы (если есть)
        if 'skipped_sheets' in abc_results and abc_results.get('skipped_sheets'):
            skipped_data = []
            for item in abc_results['skipped_sheets']:
                skipped_data.append({
                    'Лист': item.get('sheet', 'Неизвестно'),
                    'Причина пропуска': item.get('reason', 'Не указана')
                })
            
            if skipped_data:
                skipped_df = pd.DataFrame(skipped_data)
                skipped_df.to_excel(writer, sheet_name='Пропущенные_листы', index=False)
                
                worksheet = writer.sheets['Пропущенные_листы']
                worksheet.column_dimensions['A'].width = 40
                worksheet.column_dimensions['B'].width = 60
                
                wrap_alignment = Alignment(wrap_text=True, vertical='top')
                for row in range(2, len(skipped_df) + 2):
                    cell = worksheet[f'B{row}']
                    cell.alignment = wrap_alignment
    
    output.seek(0)
    return output

# Функция балансировки (без изменений)
def analyze_excel_simple(file_content):
    """Упрощенный анализ Excel файла для балансировки"""
    try:
        excel_file = pd.ExcelFile(io.BytesIO(file_content), engine='openpyxl')
        
        warehouses = []
        all_products = []
        warehouse_balances = {}
        
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
                
                if df.empty:
                    continue
                
                df.columns = [normalize_column_name(col) for col in df.columns]
                
                if 'товар' not in df.columns:
                    for col in df.columns:
                        if 'товар' in col.lower() or 'наименование' in col.lower():
                            df = df.rename(columns={col: 'товар'})
                            break
                
                if 'факт' not in df.columns:
                    for col in df.columns:
                        if 'факт' in col.lower():
                            df = df.rename(columns={col: 'факт'})
                            break
                
                if 'учет' not in df.columns:
                    for col in df.columns:
                        if 'учет' in col.lower() or 'книж' in col.lower():
                            df = df.rename(columns={col: 'учет'})
                            break
                
                if 'товар' not in df.columns or 'факт' not in df.columns or 'учет' not in df.columns:
                    continue
                
                df = df[['товар', 'факт', 'учет']].copy()
                df = df.dropna(subset=['товар'])
                df['товар'] = df['товар'].astype(str).str.strip()
                df['товар'] = df['товар'].apply(lambda x: str(x).split(' склад ')[0].strip())
                df = df[df['товар'] != '']
                
                if len(df) == 0:
                    continue
                
                df['факт_число'] = df['факт'].apply(parse_number_cached)
                df['учет_число'] = df['учет'].apply(parse_number_cached)
                df['разница'] = df['факт_число'] - df['учет_число']
                df['склад'] = sheet_name
                
                df_filtered = df[np.abs(df['разница']) > 0.001].copy()
                
                if len(df_filtered) > 0:
                    warehouses.append(sheet_name)
                    all_products.append(df_filtered)
                    
                    for _, row in df.iterrows():
                        product = row['товар']
                        if product not in warehouse_balances:
                            warehouse_balances[product] = {}
                        warehouse_balances[product][sheet_name] = row['разница']
                
            except Exception as e:
                logger.error(f"Ошибка в листе {sheet_name}: {str(e)}")
                continue
        
        if not all_products:
            return None
        
        all_data = pd.concat(all_products, ignore_index=True)
        
        movements = []
        for product, balances in warehouse_balances.items():
            if len(balances) < 2:
                continue
            
            surplus_warehouses = []
            shortage_warehouses = []
            
            for wh_name, balance in balances.items():
                if balance > 0.001:
                    surplus_warehouses.append({'склад': wh_name, 'излишек': balance})
                elif balance < -0.001:
                    shortage_warehouses.append({'склад': wh_name, 'недостача': abs(balance)})
            
            surplus_warehouses.sort(key=lambda x: x['излишек'], reverse=True)
            shortage_warehouses.sort(key=lambda x: x['недостача'], reverse=True)
            
            surplus_idx = 0
            shortage_idx = 0
            
            while surplus_idx < len(surplus_warehouses) and shortage_idx < len(shortage_warehouses):
                from_wh = surplus_warehouses[surplus_idx]
                to_wh = shortage_warehouses[shortage_idx]
                
                if from_wh['склад'] == to_wh['склад']:
                    shortage_idx += 1
                    continue
                
                amount = min(from_wh['излишек'], to_wh['недостача'])
                
                if amount > 0.001:
                    movements.append({
                        'Товар': product,
                        'Со_склада': from_wh['склад'],
                        'На_склад': to_wh['склад'],
                        'Количество': round(amount, 3)
                    })
                    
                    from_wh['излишек'] -= amount
                    to_wh['недостача'] -= amount
                    
                    if from_wh['излишек'] < 0.001:
                        surplus_idx += 1
                    if to_wh['недостача'] < 0.001:
                        shortage_idx += 1
                else:
                    break
        
        movements.sort(key=lambda x: x['Количество'], reverse=True)
        
        unique_movements = []
        seen = set()
        for move in movements:
            key = (move['Товар'], move['Со_склада'], move['На_склад'])
            if key not in seen and move['Со_склада'] != move['На_склад']:
                seen.add(key)
                unique_movements.append(move)
        
        result = {
            'warehouses': warehouses,
            'total_products': len(all_data),
            'unique_products': len(all_data['товар'].unique()),
            'total_surplus': all_data[all_data['разница'] > 0]['разница'].sum(),
            'total_shortage': abs(all_data[all_data['разница'] < 0]['разница'].sum()),
            'movements': unique_movements,
            'all_data': all_data
        }
        
        NUMBER_CACHE.clear()
        return result
        
    except Exception as e:
        logger.error(f"Ошибка анализа: {str(e)}")
        NUMBER_CACHE.clear()
        return None

# Обработчики Telegram бота
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    welcome_text = """
🏭 *БОТ ДЛЯ АНАЛИЗА СКЛАДОВ И ПРОДАЖ*

📤 *Выберите тип анализа:*
"""
    
    keyboard = [
        [InlineKeyboardButton("📊 БАЛАНСИРОВКА СКЛАДОВ", callback_data='balance')],
        [InlineKeyboardButton("📈 ABC-АНАЛИЗ ПРОДАЖ", callback_data='abc_analysis')],
        [InlineKeyboardButton("📋 Формат файлов", callback_data='format')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=welcome_text,
            parse_mode='Markdown',
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            welcome_text,
            parse_mode='Markdown',
            reply_markup=reply_markup
        )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопок"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'balance':
        balance_text = """
📊 *БАЛАНСИРОВКА СКЛАДОВ*

*Отправьте Excel файл для анализа расхождений.*

*Формат файла:*
• Каждый лист = отдельный склад
• Название листа = название склада
• Обязательные столбцы: *Товар*, *Факт*, *Учет*
"""
        keyboard = [
            [InlineKeyboardButton("📤 Отправить файл", callback_data='upload_balance')],
            [InlineKeyboardButton("🏠 Вернуться к началу", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=balance_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'abc_analysis':
        abc_text = """
📈 *ABC-АНАЛИЗ ПРОДАЖ*

*Отправьте Excel файл с данными по продажам.*

✅ *ИСПРАВЛЕНО:*
• Правильное определение столбца количества продаж
• Автоматическая настройка ширины таблиц
• Улучшено цветовое кодирование категорий
• Исправлены ошибки в рекомендациях

*Формат файла:*
• Каждый лист = отдельная точка продаж
• Автоматически определяются столбцы:
  - *Товар* (любое название)
  - *Прибыль* (столбец с числовыми значениями)
  - *Количество* (автоопределение улучшено)
"""
        keyboard = [
            [InlineKeyboardButton("📤 Отправить файл", callback_data='upload_abc')],
            [InlineKeyboardButton("🏠 Вернуться к началу", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=abc_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'upload_balance':
        await query.edit_message_text("📤 *Отправьте Excel файл для балансировки складов*", parse_mode='Markdown')
        context.user_data['analysis_type'] = 'balance'
    
    elif query.data == 'upload_abc':
        await query.edit_message_text("📤 *Отправьте Excel файл для ABC-анализа*", parse_mode='Markdown')
        context.user_data['analysis_type'] = 'abc'
    
    elif query.data == 'format':
        format_text = """
📋 *ФОРМАТ ФАЙЛОВ*

*1. Балансировка:*
   • Лист = склад
   • Столбцы: Товар, Факт, Учет

*2. ABC-анализ:*
   • Лист = точка продаж
   • Столбцы автоматически определяются
   • Поддерживает различные названия столбцов
   • Автоопределение количества продаж
"""
        keyboard = [
            [InlineKeyboardButton("📊 Балансировка", callback_data='balance')],
            [InlineKeyboardButton("📈 ABC-анализ", callback_data='abc_analysis')],
            [InlineKeyboardButton("🏠 Вернуться", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=format_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'main_menu':
        await start(update, context)

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик документов"""
    try:
        document = update.message.document
        
        if not document.file_name.lower().endswith(('.xlsx', '.xls')):
            await update.message.reply_text("❌ *Нужен Excel файл* (.xlsx или .xls)")
            return
        
        analysis_type = context.user_data.get('analysis_type')
        
        if not analysis_type:
            keyboard = [
                [InlineKeyboardButton("📊 БАЛАНСИРОВКА", callback_data='analyze_balance')],
                [InlineKeyboardButton("📈 ABC-АНАЛИЗ", callback_data='analyze_abc')],
                [InlineKeyboardButton("🏠 Вернуться к началу", callback_data='main_menu')]
            ]
            
            context.user_data['pending_file'] = {
                'file_id': document.file_id,
                'file_name': document.file_name,
                'chat_id': update.message.chat_id
            }
            
            await update.message.reply_text(
                f"📥 *Получил файл:* {document.file_name}\n\n*Выберите тип анализа:*",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return
        
        await process_document_analysis(update, context, document, analysis_type)
        
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        await update.message.reply_text(f"❌ *Ошибка:* {str(e)[:100]}")

async def process_document_analysis(update: Update, context: ContextTypes.DEFAULT_TYPE, document, analysis_type):
    """Обработка анализа"""
    try:
        chat_id = update.message.chat_id if hasattr(update, 'message') else update.callback_query.message.chat_id
        
        processing_msg = await context.bot.send_message(
            chat_id=chat_id,
            text=f"📥 *Файл:* {document.file_name}\n⏳ *Начинаю анализ...*",
            parse_mode='Markdown'
        )
        
        file = await context.bot.get_file(document.file_id)
        file_content = await file.download_as_bytearray()
        
        if analysis_type == 'balance':
            await processing_msg.edit_text(f"📥 *Файл:* {document.file_name}\n⏳ *Анализирую для балансировки...*", parse_mode='Markdown')
            result = analyze_excel_simple(file_content)
            
            if result is None:
                await processing_msg.edit_text("❌ *Не удалось проанализировать файл*")
                return
            
            # Отправка отчета балансировки
            excel_file = create_excel_report_simple(result, document.file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            await context.bot.send_document(
                chat_id=chat_id,
                document=io.BytesIO(excel_file.getvalue()),
                filename=f"Отчет_балансировка_{timestamp}.xlsx",
                caption="📊 *Отчет по балансировке складов*"
            )
            
            # Кнопки для дальнейших действий
            keyboard = [
                [InlineKeyboardButton("📤 Анализировать другой файл", callback_data='main_menu')],
                [InlineKeyboardButton("📊 Новая балансировка", callback_data='balance')],
                [InlineKeyboardButton("📈 ABC-анализ", callback_data='abc_analysis')]
            ]
            
            await context.bot.send_message(
                chat_id=chat_id,
                text="✅ *Анализ завершен!*\n\nЧто дальше?",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            await processing_msg.delete()
            
        elif analysis_type == 'abc':
            await processing_msg.edit_text(f"📥 *Файл:* {document.file_name}\n⏳ *Выполняю ABC-анализ...*", parse_mode='Markdown')
            abc_results = perform_abc_analysis(file_content)
            
            if abc_results is None:
                await processing_msg.edit_text("❌ *Не удалось выполнить ABC-анализ*")
                return
            
            # Создаем отчет
            excel_file = create_abc_excel_report(abc_results, document.file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Отправляем отчет
            await context.bot.send_document(
                chat_id=chat_id,
                document=io.BytesIO(excel_file.getvalue()),
                filename=f"Отчет_ABC_анализ_{timestamp}.xlsx",
                caption="📈 *Отчет ABC-анализа с рекомендациями*"
            )
            
            # Кнопки для дальнейших действий
            keyboard = [
                [InlineKeyboardButton("📤 Анализировать другой файл", callback_data='main_menu')],
                [InlineKeyboardButton("📊 Балансировка", callback_data='balance')],
                [InlineKeyboardButton("📈 Новый ABC-анализ", callback_data='abc_analysis')]
            ]
            
            await context.bot.send_message(
                chat_id=chat_id,
                text="✅ *ABC-анализ завершен!*\n\nЧто дальше?",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            await processing_msg.delete()
            
            # Дополнительная информация
            total_sheets = len(abc_results)
            for warehouse, data in abc_results.items():
                if 'skipped_sheets' in data:
                    skipped = len(data['skipped_sheets'])
                    if skipped > 0:
                        await context.bot.send_message(
                            chat_id=chat_id,
                            text=f"⚠️ *Внимание:* Пропущено {skipped} листов. Подробности в отчете.",
                            parse_mode='Markdown'
                        )
                        break
        
        # Очистка данных
        if 'analysis_type' in context.user_data:
            del context.user_data['analysis_type']
        if 'pending_file' in context.user_data:
            del context.user_data['pending_file']
            
    except Exception as e:
        logger.error(f"Ошибка обработки: {str(e)}")
        await context.bot.send_message(chat_id=chat_id, text=f"❌ *Ошибка:* {str(e)[:200]}")

async def analyze_type_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик выбора типа анализа"""
    query = update.callback_query
    await query.answer()
    
    pending_file = context.user_data.get('pending_file')
    if not pending_file:
        await query.edit_message_text("❌ *Ошибка:* информация о файле утеряна")
        return
    
    class MockDocument:
        def __init__(self, file_info):
            self.file_id = file_info['file_id']
            self.file_name = file_info['file_name']
    
    document = MockDocument(pending_file)
    
    if query.data == 'analyze_balance':
        context.user_data['analysis_type'] = 'balance'
        del context.user_data['pending_file']
        await query.edit_message_text(f"🔄 *Выбрана балансировка*\n📂 *Файл:* {document.file_name}", parse_mode='Markdown')
        await process_document_analysis(update, context, document, 'balance')
    
    elif query.data == 'analyze_abc':
        context.user_data['analysis_type'] = 'abc'
        del context.user_data['pending_file']
        await query.edit_message_text(f"🔄 *Выбран ABC-анализ*\n📂 *Файл:* {document.file_name}", parse_mode='Markdown')
        await process_document_analysis(update, context, document, 'abc')

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текста"""
    text = update.message.text.lower()
    
    if text in ['меню', 'start', '/start', 'начать', 'старт']:
        await start(update, context)
    elif text in ['помощь', 'help', 'справка']:
        await update.message.reply_text("👋 Используйте /start для начала работы")
    else:
        keyboard = [
            [InlineKeyboardButton("🏠 Главное меню", callback_data='main_menu')]
        ]
        await update.message.reply_text(
            "👋 *Не понимаю команду*\n\nИспользуйте кнопки ниже или команду /start",
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

def create_excel_report_simple(result, original_filename):
    """Создает отчет для балансировки с исправленными размерами ячеек"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Сводка
        summary_data = pd.DataFrame({
            'Параметр': ['Дата', 'Файл', 'Складов', 'Товаров', 'Излишек', 'Недостача', 'Перемещений'],
            'Значение': [
                datetime.now().strftime('%d.%m.%Y %H:%M'),
                original_filename,
                len(result['warehouses']),
                result['total_products'],
                f"{result['total_surplus']:.3f}",
                f"{result['total_shortage']:.3f}",
                len(result['movements'])
            ]
        })
        summary_data.to_excel(writer, sheet_name='Сводка', index=False)
        
        # Настраиваем ширину столбцов для сводки
        worksheet = writer.sheets['Сводка']
        worksheet.column_dimensions['A'].width = 25  # Параметр
        worksheet.column_dimensions['B'].width = 40  # Значение
        
        # Перемещения
        if result['movements']:
            moves_data = []
            for move in result['movements']:
                moves_data.append({
                    'Товар': move['Товар'][:50],
                    'Со склада': move['Со_склада'],
                    'На склад': move['На_склад'],
                    'Количество': move['Количество']
                })
            
            moves_df = pd.DataFrame(moves_data)
            moves_df.to_excel(writer, sheet_name='Перемещения', index=False)
            
            # Настраиваем ширину столбцов для перемещений
            worksheet = writer.sheets['Перемещения']
            worksheet.column_dimensions['A'].width = 40  # Товар
            worksheet.column_dimensions['B'].width = 25  # Со склада
            worksheet.column_dimensions['C'].width = 25  # На склад
            worksheet.column_dimensions['D'].width = 15  # Количество
    
    output.seek(0)
    return output

def main():
    """Запуск бота"""
    try:
        print("=" * 60)
        print("🚀 БОТ ДЛЯ АНАЛИЗА СКЛАДОВ И ПРОДАЖ")
        print("=" * 60)
        print("✅ Версия: 4.7 (Исправлены ошибки определения количества)")
        print("✅ Исправлено:")
        print("   • Улучшено определение столбца количества продаж")
        print("   • Автоматическая настройка ширины столбцов")
        print("   • Исправлены ошибки в рекомендациях")
        print("   • Улучшено цветовое кодирование категорий ABC")
        print("=" * 60)
        
        application = Application.builder().token(TOKEN).build()
        
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CallbackQueryHandler(button_handler))
        application.add_handler(CallbackQueryHandler(analyze_type_handler, pattern='^analyze_'))
        application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
        
        print("✅ Бот запущен!")
        print("=" * 60)
        
        application.run_polling(allowed_updates=Update.ALL_TYPES)
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()