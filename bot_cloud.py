import pandas as pd
import numpy as np
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
import asyncio
import io
import re
from collections import defaultdict
import warnings
from datetime import datetime
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import signal
import sys
import pytz

warnings.filterwarnings('ignore')

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Конфигурация бота
TOKEN = '8234604599:AAFluxkjpDxUDz2kgvNYTwGIkMi5NuWrfmU'

# Константы для ABC-анализа
ABC_CATEGORIES = {
    'A': {'min_percent': 0, 'max_percent': 80, 'description': 'Критически важные товары'},
    'B': {'min_percent': 80, 'max_percent': 95, 'description': 'Средняя значимость'},
    'C': {'min_percent': 95, 'max_percent': 100, 'description': 'Наименьшая значимость'}
}

# Цвета для категорий ABC
ABC_COLORS = {
    'A': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'B': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'C': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

# Шрифт для категорий
CATEGORY_FONT = Font(color='000000', bold=True)

# Маппинги для столбцов
COLUMN_MAPPINGS = {
    'товар': ['товар', 'наименование', 'название', 'name', 'product', 'артикул', 'код'],
    'факт': ['факт', 'фактическое', 'факт.', 'actual', 'fact', 'кол-во'],
    'учет': ['учет', 'книжн', 'книжное', 'бухг', 'учетное', 'book'],
    'разница': ['разница', 'дельта', 'diff', 'difference'],
    'количество': ['количество', 'quantity', 'qty', 'кол-во', 'продажи', 'ед.'],
    'выручка': ['выручка', 'revenue', 'продажи', 'sales', 'выручка, р.'],
    'прибыль': ['прибыль', 'profit', 'маржа', 'валовая прибыль'],
    'наценка': ['наценка', 'markup', 'рентабельность']
}

def normalize_column_name(name):
    """Нормализация названий столбцов"""
    if not isinstance(name, str):
        name = str(name)
    
    name_lower = name.lower().strip()
    
    for std_name, variants in COLUMN_MAPPINGS.items():
        for variant in variants:
            if variant in name_lower:
                return std_name
    return name_lower

NUMBER_CACHE = {}

def parse_number_cached(value):
    """Кешированная функция парсинга чисел"""
    if pd.isna(value):
        return 0.0
    
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    
    if isinstance(value, str):
        if value in NUMBER_CACHE:
            return NUMBER_CACHE[value]
        
        val = value.strip()
        if not val:
            result = 0.0
        else:
            val = val.replace(',', '.').replace(' ', '')
            val = re.sub(r'[^\d\.\-]', '', val)
            
            try:
                result = float(val)
            except ValueError:
                result = 0.0
        
        NUMBER_CACHE[value] = result
        return result
    
    return 0.0

# Функция ABC-анализа
def perform_abc_analysis(file_content):
    """Выполняет ABC-анализ по файлу продаж"""
    try:
        excel_file = pd.ExcelFile(io.BytesIO(file_content), engine='openpyxl')
        
        all_abc_results = {}
        processed_sheets = []
        skipped_sheets = []
        
        for sheet_name in excel_file.sheet_names:
            try:
                logger.info(f"Обработка листа: {sheet_name}")
                
                df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
                
                if df.empty:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Пустой лист'})
                    continue
                
                df.columns = [normalize_column_name(col) for col in df.columns]
                
                if 'товар' not in df.columns:
                    for col in df.columns:
                        sample_values = df[col].dropna().head(10).astype(str).tolist()
                        sample_str = ' '.join(sample_values).lower()
                        if any(keyword in sample_str for keyword in ['пицца', 'бургер', 'латте', 'кофе', 'салат', 'суп']):
                            df = df.rename(columns={col: 'товар'})
                            break
                
                if 'товар' not in df.columns:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Не найден столбец товара'})
                    continue
                
                profit_column = None
                quantity_column = None
                
                for col in df.columns:
                    col_lower = col.lower()
                    if 'прибыль' in col_lower or 'profit' in col_lower:
                        profit_column = col
                    elif 'количество' in col_lower or 'qty' in col_lower or 'кол-во' in col_lower:
                        quantity_column = col
                
                if not profit_column:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Не найден столбец прибыли'})
                    continue
                
                columns_to_use = ['товар', profit_column]
                if quantity_column:
                    columns_to_use.append(quantity_column)
                
                df_clean = df[columns_to_use].copy()
                df_clean = df_clean.dropna(subset=['товар'])
                df_clean['товар'] = df_clean['товар'].astype(str).str.strip()
                
                summary_keywords = ['итого', 'total', 'всего', 'сумма']
                df_clean = df_clean[~df_clean['товар'].str.lower().isin([kw.lower() for kw in summary_keywords])]
                df_clean = df_clean[df_clean['товар'] != '']
                
                if len(df_clean) == 0:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Нет данных после очистки'})
                    continue
                
                df_clean['прибыль_число'] = df_clean[profit_column].apply(parse_number_cached)
                df_clean = df_clean[df_clean['прибыль_число'] > 0]
                
                if len(df_clean) == 0:
                    skipped_sheets.append({'sheet': sheet_name, 'reason': 'Нет положительной прибыли'})
                    continue
                
                if quantity_column:
                    df_clean['количество_число'] = df_clean[quantity_column].apply(parse_number_cached)
                else:
                    df_clean['количество_число'] = 1
                
                # ABC-анализ
                df_sorted_profit = df_clean.sort_values('прибыль_число', ascending=False).reset_index(drop=True)
                total_profit = df_sorted_profit['прибыль_число'].sum()
                
                df_sorted_profit['доля_в_прибыли'] = (df_sorted_profit['прибыль_число'] / total_profit) * 100
                df_sorted_profit['кумулятивная_доля_прибыли'] = df_sorted_profit['доля_в_прибыли'].cumsum()
                
                def assign_abc_category(cumulative_share):
                    if cumulative_share <= 80:
                        return 'A'
                    elif cumulative_share <= 95:
                        return 'B'
                    else:
                        return 'C'
                
                df_sorted_profit['категория_abc_прибыль'] = df_sorted_profit['кумулятивная_доля_прибыли'].apply(assign_abc_category)
                
                total_qty = df_sorted_profit['количество_число'].sum()
                df_sorted_profit['доля_в_количестве'] = (df_sorted_profit['количество_число'] / total_qty) * 100
                df_sorted_profit['кумулятивная_доля_количества'] = df_sorted_profit['доля_в_количестве'].cumsum()
                df_sorted_profit['категория_abc_количество'] = df_sorted_profit['кумулятивная_доля_количества'].apply(assign_abc_category)
                
                df_sorted_profit['позиция_прибыль'] = df_sorted_profit.index + 1
                
                category_stats = {}
                for category in ['A', 'B', 'C']:
                    cat_data = df_sorted_profit[df_sorted_profit['категория_abc_прибыль'] == category]
                    if len(cat_data) > 0:
                        category_stats[category] = {
                            'количество_товаров': len(cat_data),
                            'доля_товаров': (len(cat_data) / len(df_sorted_profit)) * 100,
                            'сумма_прибыли': cat_data['прибыль_число'].sum(),
                            'доля_прибыли': cat_data['доля_в_прибыли'].sum(),
                            'средняя_прибыль': cat_data['прибыль_число'].mean(),
                            'сумма_количества': cat_data['количество_число'].sum(),
                            'среднее_количество': cat_data['количество_число'].mean()
                        }
                    else:
                        category_stats[category] = {
                            'количество_товаров': 0,
                            'доля_товаров': 0,
                            'сумма_прибыли': 0,
                            'доля_прибыли': 0,
                            'средняя_прибыль': 0,
                            'сумма_количества': 0,
                            'среднее_количество': 0
                        }
                
                top_a_products = df_sorted_profit[df_sorted_profit['категория_abc_прибыль'] == 'A'][['товар', 'прибыль_число', 'доля_в_прибыли', 'количество_число']].head(10)
                
                all_abc_results[sheet_name] = {
                    'dataframe': df_sorted_profit,
                    'total_profit': total_profit,
                    'total_quantity': total_qty,
                    'total_products': len(df_sorted_profit),
                    'category_stats': category_stats,
                    'top_a_products': top_a_products,
                    'profit_column': profit_column,
                    'quantity_column': quantity_column if quantity_column else 'не найден'
                }
                
                processed_sheets.append(sheet_name)
                
            except Exception as e:
                skipped_sheets.append({'sheet': sheet_name, 'reason': f'Ошибка: {str(e)[:100]}'})
                continue
        
        NUMBER_CACHE.clear()
        
        if all_abc_results:
            for sheet_name in all_abc_results.keys():
                all_abc_results[sheet_name]['processed_sheets'] = processed_sheets
                all_abc_results[sheet_name]['skipped_sheets'] = skipped_sheets
        
        return all_abc_results if all_abc_results else None
        
    except Exception as e:
        logger.error(f"Ошибка ABC-анализа: {str(e)}")
        NUMBER_CACHE.clear()
        return None

# Функция создания отчета
def create_abc_excel_report(abc_results, original_filename):
    """Создает Excel файл с отчетами"""
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Сводный отчет
        summary_data = []
        for warehouse, data in abc_results.items():
            stats = data['category_stats']
            summary_data.append({
                'Склад/Точка': warehouse,
                'Всего товаров': data['total_products'],
                'Общая прибыль': data['total_profit'],
                'Общее количество': data['total_quantity'],
                'Товаров A': stats['A']['количество_товаров'],
                'Доля A, %': f"{stats['A']['доля_прибыли']:.1f}%",
                'Товаров B': stats['B']['количество_товаров'],
                'Доля B, %': f"{stats['B']['доля_прибыли']:.1f}%",
                'Товаров C': stats['C']['количество_товаров'],
                'Доля C, %': f"{stats['C']['доля_прибыли']:.1f}%"
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df['Общая прибыль'] = summary_df['Общая прибыль'].apply(lambda x: f"{x:,.2f} р.")
        summary_df['Общее количество'] = summary_df['Общее количество'].apply(lambda x: f"{x:,.0f}")
        summary_df.to_excel(writer, sheet_name='Сводка', index=False)
        
        # Детальный анализ
        for warehouse, data in abc_results.items():
            df = data['dataframe'].copy()
            
            report_df = df[[
                'позиция_прибыль', 'товар', 'прибыль_число', 'доля_в_прибыли', 
                'кумулятивная_доля_прибыли', 'категория_abc_прибыль',
                'количество_число', 'доля_в_количестве', 'кумулятивная_доля_количества', 'категория_abc_количество'
            ]].copy()
            
            report_df.columns = [
                'Позиция', 'Товар', 'Прибыль, руб.', 'Доля в прибыли, %',
                'Кумулятивная доля прибыли, %', 'Категория ABC (прибыль)',
                'Количество продаж', 'Доля в количестве, %',
                'Кумулятивная доля количества, %', 'Категория ABC (количество)'
            ]
            
            report_df['Прибыль, руб.'] = report_df['Прибыль, руб.'].apply(lambda x: f"{x:,.2f}")
            report_df['Доля в прибыли, %'] = report_df['Доля в прибыли, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Кумулятивная доля прибыли, %'] = report_df['Кумулятивная доля прибыли, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Количество продаж'] = report_df['Количество продаж'].apply(lambda x: f"{x:,.0f}")
            report_df['Доля в количестве, %'] = report_df['Доля в количестве, %'].apply(lambda x: f"{x:.2f}%")
            report_df['Кумулятивная доля количества, %'] = report_df['Кумулятивная доля количества, %'].apply(lambda x: f"{x:.2f}%")
            
            sheet_name = f"ABC_{warehouse[:20]}" if len(warehouse) > 20 else f"ABC_{warehouse}"
            sheet_name = re.sub(r'[\\/*?:\[\]]', '', sheet_name)
            
            report_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            if sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                for row in range(2, len(report_df) + 2):
                    profit_cell = worksheet[f'F{row}']
                    category = profit_cell.value
                    if category in ABC_COLORS:
                        profit_cell.fill = ABC_COLORS[category]
                        profit_cell.font = CATEGORY_FONT
                        profit_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    qty_cell = worksheet[f'J{row}']
                    category_qty = qty_cell.value
                    if category_qty in ABC_COLORS:
                        qty_cell.fill = ABC_COLORS[category_qty]
                        qty_cell.font = CATEGORY_FONT
                        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    output.seek(0)
    return output

# Функция балансировки
def analyze_excel_simple(file_content):
    """Упрощенный анализ Excel файла для балансировки"""
    try:
        excel_file = pd.ExcelFile(io.BytesIO(file_content), engine='openpyxl')
        
        warehouses = []
        all_products = []
        warehouse_balances = {}
        
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str)
                
                if df.empty:
                    continue
                
                df.columns = [normalize_column_name(col) for col in df.columns]
                
                if 'товар' not in df.columns:
                    for col in df.columns:
                        if 'товар' in col.lower():
                            df = df.rename(columns={col: 'товар'})
                            break
                
                if 'факт' not in df.columns:
                    for col in df.columns:
                        if 'факт' in col.lower():
                            df = df.rename(columns={col: 'факт'})
                            break
                
                if 'учет' not in df.columns:
                    for col in df.columns:
                        if 'учет' in col.lower():
                            df = df.rename(columns={col: 'учет'})
                            break
                
                if 'товар' not in df.columns or 'факт' not in df.columns or 'учет' not in df.columns:
                    continue
                
                df = df[['товар', 'факт', 'учет']].copy()
                df = df.dropna(subset=['товар'])
                df['товар'] = df['товар'].astype(str).str.strip()
                df = df[df['товар'] != '']
                
                if len(df) == 0:
                    continue
                
                df['факт_число'] = df['факт'].apply(parse_number_cached)
                df['учет_число'] = df['учет'].apply(parse_number_cached)
                df['разница'] = df['факт_число'] - df['учет_число']
                df['склад'] = sheet_name
                
                df_filtered = df[np.abs(df['разница']) > 0.001].copy()
                
                if len(df_filtered) > 0:
                    warehouses.append(sheet_name)
                    all_products.append(df_filtered)
                    
                    for _, row in df.iterrows():
                        product = row['товар']
                        if product not in warehouse_balances:
                            warehouse_balances[product] = {}
                        warehouse_balances[product][sheet_name] = row['разница']
                
            except Exception as e:
                continue
        
        if not all_products:
            return None
        
        all_data = pd.concat(all_products, ignore_index=True)
        
        movements = []
        for product, balances in warehouse_balances.items():
            if len(balances) < 2:
                continue
            
            surplus_warehouses = []
            shortage_warehouses = []
            
            for wh_name, balance in balances.items():
                if balance > 0.001:
                    surplus_warehouses.append({'склад': wh_name, 'излишек': balance})
                elif balance < -0.001:
                    shortage_warehouses.append({'склад': wh_name, 'недостача': abs(balance)})
            
            surplus_warehouses.sort(key=lambda x: x['излишек'], reverse=True)
            shortage_warehouses.sort(key=lambda x: x['недостача'], reverse=True)
            
            surplus_idx = 0
            shortage_idx = 0
            
            while surplus_idx < len(surplus_warehouses) and shortage_idx < len(shortage_warehouses):
                from_wh = surplus_warehouses[surplus_idx]
                to_wh = shortage_warehouses[shortage_idx]
                
                if from_wh['склад'] == to_wh['склад']:
                    shortage_idx += 1
                    continue
                
                amount = min(from_wh['излишек'], to_wh['недостача'])
                
                if amount > 0.001:
                    movements.append({
                        'Товар': product,
                        'Со_склада': from_wh['склад'],
                        'На_склад': to_wh['склад'],
                        'Количество': round(amount, 3)
                    })
                    
                    from_wh['излишек'] -= amount
                    to_wh['недостача'] -= amount
                    
                    if from_wh['излишек'] < 0.001:
                        surplus_idx += 1
                    if to_wh['недостача'] < 0.001:
                        shortage_idx += 1
                else:
                    break
        
        movements.sort(key=lambda x: x['Количество'], reverse=True)
        
        unique_movements = []
        seen = set()
        for move in movements:
            key = (move['Товар'], move['Со_склада'], move['На_склад'])
            if key not in seen and move['Со_склада'] != move['На_склад']:
                seen.add(key)
                unique_movements.append(move)
        
        result = {
            'warehouses': warehouses,
            'total_products': len(all_data),
            'unique_products': len(all_data['товар'].unique()),
            'total_surplus': all_data[all_data['разница'] > 0]['разница'].sum(),
            'total_shortage': abs(all_data[all_data['разница'] < 0]['разница'].sum()),
            'movements': unique_movements,
            'all_data': all_data
        }
        
        NUMBER_CACHE.clear()
        return result
        
    except Exception as e:
        NUMBER_CACHE.clear()
        return None

# Обработчики Telegram бота
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    welcome_text = """
🏭 *БОТ ДЛЯ АНАЛИЗА СКЛАДОВ И ПРОДАЖ*

📤 *Выберите тип анализа:*
"""
    
    keyboard = [
        [InlineKeyboardButton("📊 БАЛАНСИРОВКА СКЛАДОВ", callback_data='balance')],
        [InlineKeyboardButton("📈 ABC-АНАЛИЗ ПРОДАЖ", callback_data='abc_analysis')],
        [InlineKeyboardButton("ℹ️ Информация", callback_data='info')]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text=welcome_text,
            parse_mode='Markdown',
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            welcome_text,
            parse_mode='Markdown',
            reply_markup=reply_markup
        )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик кнопок"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'balance':
        balance_text = """
📊 *БАЛАНСИРОВКА СКЛАДОВ*

*Отправьте Excel файл для анализа расхождений.*

*Формат файла:*
• Каждый лист = отдельный склад
• Название листа = название склада
• Обязательные столбцы: *Товар*, *Факт*, *Учет*
"""
        keyboard = [
            [InlineKeyboardButton("📤 Отправить файл", callback_data='upload_balance')],
            [InlineKeyboardButton("🏠 Вернуться к началу", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=balance_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'abc_analysis':
        abc_text = """
📈 *ABC-АНАЛИЗ ПРОДАЖ*

*Отправьте Excel файл с данными по продажам.*

*Формат файла:*
• Каждый лист = отдельная точка продаж
• Автоматически определяются столбцы:
  - *Товар* (любое название)
  - *Прибыль* (столбец с числовыми значениями)
  - *Количество* (опционально)
"""
        keyboard = [
            [InlineKeyboardButton("📤 Отправить файл", callback_data='upload_abc')],
            [InlineKeyboardButton("🏠 Вернуться к началу", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=abc_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'upload_balance':
        await query.edit_message_text("📤 *Отправьте Excel файл для балансировки складов*", parse_mode='Markdown')
        context.user_data['analysis_type'] = 'balance'
    
    elif query.data == 'upload_abc':
        await query.edit_message_text("📤 *Отправьте Excel файл для ABC-анализа*", parse_mode='Markdown')
        context.user_data['analysis_type'] = 'abc'
    
    elif query.data == 'info':
        info_text = """
🤖 *ИНФОРМАЦИЯ О БОТЕ*

*Версия:* 4.8 (Облачная)
*Статус:* Работает 24/7
*Функции:*
• Балансировка складов
• ABC-анализ продаж
• Автоматические отчеты

*Бот работает в облаке и доступен всегда!*
"""
        keyboard = [
            [InlineKeyboardButton("📊 Балансировка", callback_data='balance')],
            [InlineKeyboardButton("📈 ABC-анализ", callback_data='abc_analysis')],
            [InlineKeyboardButton("🏠 Главное меню", callback_data='main_menu')]
        ]
        await query.edit_message_text(text=info_text, parse_mode='Markdown', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif query.data == 'main_menu':
        await start(update, context)

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик документов"""
    try:
        document = update.message.document
        
        if not document.file_name.lower().endswith(('.xlsx', '.xls')):
            await update.message.reply_text("❌ *Нужен Excel файл* (.xlsx или .xls)")
            return
        
        analysis_type = context.user_data.get('analysis_type')
        
        if not analysis_type:
            keyboard = [
                [InlineKeyboardButton("📊 БАЛАНСИРОВКА", callback_data='analyze_balance')],
                [InlineKeyboardButton("📈 ABC-АНАЛИЗ", callback_data='analyze_abc')],
                [InlineKeyboardButton("🏠 Главное меню", callback_data='main_menu')]
            ]
            
            context.user_data['pending_file'] = {
                'file_id': document.file_id,
                'file_name': document.file_name,
                'chat_id': update.message.chat_id
            }
            
            await update.message.reply_text(
                f"📥 *Получил файл:* {document.file_name}\n\n*Выберите тип анализа:*",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return
        
        await process_document_analysis(update, context, document, analysis_type)
        
    except Exception as e:
        await update.message.reply_text(f"❌ *Ошибка:* {str(e)[:100]}")

async def process_document_analysis(update: Update, context: ContextTypes.DEFAULT_TYPE, document, analysis_type):
    """Обработка анализа"""
    try:
        chat_id = update.message.chat_id
        
        processing_msg = await context.bot.send_message(
            chat_id=chat_id,
            text=f"📥 *Файл:* {document.file_name}\n⏳ *Начинаю анализ...*",
            parse_mode='Markdown'
        )
        
        file = await context.bot.get_file(document.file_id)
        file_content = await file.download_as_bytearray()
        
        if analysis_type == 'balance':
            await processing_msg.edit_text(f"📥 *Файл:* {document.file_name}\n⏳ *Анализирую для балансировки...*", parse_mode='Markdown')
            result = analyze_excel_simple(file_content)
            
            if result is None:
                await processing_msg.edit_text("❌ *Не удалось проанализировать файл*")
                return
            
            excel_file = create_excel_report_simple(result, document.file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            await context.bot.send_document(
                chat_id=chat_id,
                document=io.BytesIO(excel_file.getvalue()),
                filename=f"Отчет_балансировка_{timestamp}.xlsx",
                caption="📊 *Отчет по балансировке складов*"
            )
            
            keyboard = [
                [InlineKeyboardButton("📤 Анализировать другой файл", callback_data='main_menu')],
                [InlineKeyboardButton("📊 Новая балансировка", callback_data='balance')],
                [InlineKeyboardButton("📈 ABC-анализ", callback_data='abc_analysis')]
            ]
            
            await context.bot.send_message(
                chat_id=chat_id,
                text="✅ *Анализ завершен!*\n\nЧто дальше?",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            await processing_msg.delete()
            
        elif analysis_type == 'abc':
            await processing_msg.edit_text(f"📥 *Файл:* {document.file_name}\n⏳ *Выполняю ABC-анализ...*", parse_mode='Markdown')
            abc_results = perform_abc_analysis(file_content)
            
            if abc_results is None:
                await processing_msg.edit_text("❌ *Не удалось выполнить ABC-анализ*")
                return
            
            excel_file = create_abc_excel_report(abc_results, document.file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            await context.bot.send_document(
                chat_id=chat_id,
                document=io.BytesIO(excel_file.getvalue()),
                filename=f"Отчет_ABC_анализ_{timestamp}.xlsx",
                caption="📈 *Отчет ABC-анализа с рекомендациями*"
            )
            
            keyboard = [
                [InlineKeyboardButton("📤 Анализировать другой файл", callback_data='main_menu')],
                [InlineKeyboardButton("📊 Балансировка", callback_data='balance')],
                [InlineKeyboardButton("📈 Новый ABC-анализ", callback_data='abc_analysis')]
            ]
            
            await context.bot.send_message(
                chat_id=chat_id,
                text="✅ *ABC-анализ завершен!*\n\nЧто дальше?",
                parse_mode='Markdown',
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            
            await processing_msg.delete()
        
        if 'analysis_type' in context.user_data:
            del context.user_data['analysis_type']
        if 'pending_file' in context.user_data:
            del context.user_data['pending_file']
            
    except Exception as e:
        await context.bot.send_message(chat_id=chat_id, text=f"❌ *Ошибка:* {str(e)[:200]}")

async def analyze_type_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик выбора типа анализа"""
    query = update.callback_query
    await query.answer()
    
    pending_file = context.user_data.get('pending_file')
    if not pending_file:
        await query.edit_message_text("❌ *Ошибка:* информация о файле утеряна")
        return
    
    class MockDocument:
        def __init__(self, file_info):
            self.file_id = file_info['file_id']
            self.file_name = file_info['file_name']
    
    document = MockDocument(pending_file)
    
    if query.data == 'analyze_balance':
        context.user_data['analysis_type'] = 'balance'
        del context.user_data['pending_file']
        await query.edit_message_text(f"🔄 *Выбрана балансировка*\n📂 *Файл:* {document.file_name}", parse_mode='Markdown')
        await process_document_analysis(update, context, document, 'balance')
    
    elif query.data == 'analyze_abc':
        context.user_data['analysis_type'] = 'abc'
        del context.user_data['pending_file']
        await query.edit_message_text(f"🔄 *Выбран ABC-анализ*\n📂 *Файл:* {document.file_name}", parse_mode='Markdown')
        await process_document_analysis(update, context, document, 'abc')

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик текста"""
    text = update.message.text.lower()
    
    if text in ['меню', 'start', '/start', 'начать', 'старт']:
        await start(update, context)
    elif text in ['помощь', 'help', 'справка']:
        await update.message.reply_text("👋 Используйте /start для начала работы")
    elif text == 'статус':
        await update.message.reply_text("✅ *Бот работает 24/7 в облаке!*\n\nВсе функции доступны.", parse_mode='Markdown')
    else:
        keyboard = [
            [InlineKeyboardButton("🏠 Главное меню", callback_data='main_menu')]
        ]
        await update.message.reply_text(
            "👋 *Не понимаю команду*\n\nИспользуйте кнопки ниже или команду /start",
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

def create_excel_report_simple(result, original_filename):
    """Создает отчет для балансировки"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_data = pd.DataFrame({
            'Параметр': ['Дата', 'Файл', 'Складов', 'Товаров', 'Излишек', 'Недостача', 'Перемещений'],
            'Значение': [
                datetime.now().strftime('%d.%m.%Y %H:%M'),
                original_filename,
                len(result['warehouses']),
                result['total_products'],
                f"{result['total_surplus']:.3f}",
                f"{result['total_shortage']:.3f}",
                len(result['movements'])
            ]
        })
        summary_data.to_excel(writer, sheet_name='Сводка', index=False)
        
        worksheet = writer.sheets['Сводка']
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 40
        
        if result['movements']:
            moves_data = []
            for move in result['movements']:
                moves_data.append({
                    'Товар': move['Товар'][:50],
                    'Со склада': move['Со_склада'],
                    'На склад': move['На_склад'],
                    'Количество': move['Количество']
                })
            
            moves_df = pd.DataFrame(moves_data)
            moves_df.to_excel(writer, sheet_name='Перемещения', index=False)
            
            worksheet = writer.sheets['Перемещения']
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 15
    
    output.seek(0)
    return output

def signal_handler(signum, frame):
    """Обработчик сигналов для корректного завершения"""
    print(f"\n🚦 Получен сигнал {signum}, завершаю работу...")
    print("👋 До свидания!")
    sys.exit(0)

async def send_health_check():
    """Периодическая проверка здоровья бота"""
    while True:
        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"✅ Бот работает: {current_time}")
        except Exception as e:
            print(f"❌ Ошибка health check: {e}")
        
        await asyncio.sleep(3600)  # Проверка каждый час

async def main_async():
    """Асинхронная основная функция"""
    try:
        # Регистрация обработчиков сигналов
        signal.signal(signal.SIGINT, signal_handler)
        signal.signal(signal.SIGTERM, signal_handler)
        
        print("=" * 60)
        print("🚀 БОТ ДЛЯ АНАЛИЗА СКЛАДОВ И ПРОДАЖ")
        print("=" * 60)
        print("✅ Версия: 4.8 (Облачная версия 24/7)")
        print("✅ Статус: Запускается...")
        print("=" * 60)
        
        # Создаем приложение
        application = Application.builder().token(TOKEN).build()
        
        # Регистрация обработчиков
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CallbackQueryHandler(button_handler))
        application.add_handler(CallbackQueryHandler(analyze_type_handler, pattern='^analyze_'))
        application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
        
        print("✅ Бот инициализирован")
        print("✅ Обработчики зарегистрированы")
        print("=" * 60)
        print("✅ Бот запущен и работает 24/7!")
        print("=" * 60)
        print("ℹ️  Для остановки нажмите Ctrl+C")
        print("=" * 60)
        
        # Запускаем health check в фоне
        health_task = asyncio.create_task(send_health_check())
        
        # Запускаем бота
        await application.initialize()
        await application.start()
        await application.updater.start_polling(
            allowed_updates=Update.ALL_TYPES,
            timeout=30,
            pool_timeout=30,
            connect_timeout=30,
            drop_pending_updates=True
        )
        
        # Ждем вечно (пока не получим сигнал остановки)
        await asyncio.Event().wait()
        
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("🔄 Завершение работы бота...")

def main():
    """Точка входа"""
    try:
        # Запускаем асинхронную функцию
        asyncio.run(main_async())
    except KeyboardInterrupt:
        print("\n👋 Бот остановлен пользователем")
    except Exception as e:
        print(f"❌ Ошибка запуска: {e}")

if __name__ == '__main__':
    main()